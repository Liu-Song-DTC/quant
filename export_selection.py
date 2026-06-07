"""
选股结果导出到 Excel — 用于实盘选股记录和后续换股操作

用法: python export_selection.py

设计:
  - 使用完整 dynamic 因子模式选股（factor_mode=both）
  - factor_df 在预计算后自动释放，避免 OOM
  - 5个Sheet: 本次选股 / 调仓对比 / 市场状态 / 调仓历史 / 信号排名Top50
"""
import os
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.resolve()

sys.path.insert(0, str(ROOT / "trade"))
sys.path.insert(0, str(ROOT / "strategy"))
sys.path.insert(0, str(ROOT / "data"))

from trade.config import TradeConfig


# ══════════════════════════════════════════════════════════════
#  Excel 样式工具
# ══════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
EVEN_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")


def style_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, ncols):
    """批量设置数据行样式 — 使用 iter_rows 替代逐格 cell() 查找"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row,
                                                 min_col=1, max_col=ncols)):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for cell in row:
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
            if cell.fill == PatternFill():  # 不覆盖已设置的填充色
                cell.fill = fill


def auto_width(ws, min_width=8, max_width=40):
    """计算列宽 — 单次遍历 iter_rows 替代 ws.columns（openpyxl ws.columns 极慢）"""
    col_widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value:
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                col_idx = cell.col_idx
                if length > col_widths.get(col_idx, 0):
                    col_widths[col_idx] = length
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(width + 2, min_width), max_width)


def load_stock_name_map():
    stock_list = ROOT / "data" / "stock_data" / "stock_metadata" / "stock_list.csv"
    if stock_list.exists():
        df = pd.read_csv(stock_list, dtype={"symbol": str}, usecols=["symbol", "name"])
        return dict(zip(df["symbol"], df["name"]))
    return {}


# ══════════════════════════════════════════════════════════════
#  选股理由生成 — 缠论逐步推理 + 课文引用
#  面向他人阅读的报告格式：假设→验证→结论
# ══════════════════════════════════════════════════════════════

# 缠论买卖点 → 课文出处与推理链条
CHAN_BUY_REASONING = {
    1: {
        'lesson': '第21课',
        'premise': '下跌趋势末端出现底背驰 → 空头力竭',
        'verification': '底背驰强度{strength}，MACD柱面积缩小(第24课)',
        'conclusion': '符合第一类买点定义：趋势背驰终结下跌走势',
    },
    2: {
        'lesson': '第21/49课',
        'premise': '一买后回踩不破前低 → 空头无力继续打压',
        'verification': '回踩确认强度{strength}，第49课"最安全的买点"',
        'conclusion': '符合第二类买点定义：次级别回踩确认',
    },
    3: {
        'lesson': '第21/54课',
        'premise': '突破中枢上沿后回踩不破 → 多头主导',
        'verification': '中枢突破力度{strength}，第54课"走势加速确认"',
        'conclusion': '符合第三类买点定义：中枢突破后回踩确认',
    },
}


def generate_selection_reason(s, rank_pct_str=""):
    """
    面向报告的逐步推理格式 (用于CSV历史记录，简明版)
    """
    buy_pt = s.get("chan_buy_point", 0)
    buy_strength = s.get("chan_buy_strength", 0.0)
    signal_level = s.get("signal_level", 0)
    trend = s.get("trend_type", 0)
    chan_bonus = s.get("chan_bonus", 0)
    fn = s.get("factor_name", "")

    if buy_pt <= 0:
        return f'因子评分驱动·无缠论买点确认, 综合评分{s.get("score", 0):.3f}'

    reasoning = CHAN_BUY_REASONING.get(buy_pt, {})
    lesson = reasoning.get('lesson', '')
    premise = reasoning.get('premise', '')

    # 趋势背景 (第15课:没有趋势没有背驰)
    trend_words = {2: '上涨趋势延续·顺势做多(第16课)', 1: '盘整中·等待方向(第18课)', -2: '下跌末端·左侧转折(第21课)', 0: ''}
    trend_str = trend_words.get(trend, '')

    # 信号级别 (第35课:多级别联立)
    lvl_words = {3: '笔线段双级共振(第35课)', 2: '线段级确认(第35课)', 1: '', 0: ''}
    lvl_str = lvl_words.get(signal_level, '')

    parts = [premise]
    if trend_str:
        parts.append(trend_str)
    if lvl_str:
        parts.append(lvl_str)
    if chan_bonus >= 0.06:
        parts.append('缠论增强')
    if fn and fn.startswith('DYN_'):
        parts.append('动态IC因子验证有效')
    if rank_pct_str:
        parts.append(f'截面{rank_pct_str}')

    return f'{lesson}: ' + '; '.join(parts)


def generate_selection_report(s, rank_pct_str="", stock_name=""):
    """
    生成完整的选股分析段落 — 面向他人阅读的报告格式
    结构: 背景 → 缠论逐步分析 → 验证 → 结论
    """
    buy_pt = s.get("chan_buy_point", 0)
    buy_strength = s.get("chan_buy_strength", 0.0)
    signal_level = s.get("signal_level", 0)
    trend = s.get("trend_type", 0)
    chan_bonus = s.get("chan_bonus", 0)
    fn = s.get("factor_name", "")
    score = s.get("score", 0)
    code = s.get("code", "")
    industry = s.get("industry", "")
    name = stock_name or code

    if buy_pt <= 0:
        return (f'{name}({code}) 以综合评分{score:.3f}入选，'
                f'主要依赖因子面驱动，无明确缠论买点确认。'
                f'建议关注后续走势结构演化。')

    reasoning = CHAN_BUY_REASONING.get(buy_pt, {})
    lesson = reasoning.get('lesson', '')
    premise = reasoning.get('premise', '')
    verification_tpl = reasoning.get('verification', '')
    conclusion = reasoning.get('conclusion', '')

    verification = verification_tpl.format(strength=f'{buy_strength:.2f}')

    # 趋势
    trend_analysis = {2: '个股处于上涨趋势中，EMA多头排列(第11课)，顺势环境有利于买点成功。',
                      1: '个股处于盘整状态，中枢震荡中(第18课)，需关注突破方向。',
                      -2: '个股处于下跌趋势末端，走势终完美(第17课)，转折信号需进一步确认。',
                      0: '趋势尚不明确，需等待方向选择。'}.get(trend, '')

    # 信号级别
    lvl_analysis = {3: '笔级别与线段级别同时发出买入信号(第35课)，双周期共振显著提高可靠性。',
                    2: '信号已获线段级别确认(第35课)，中期趋势转折概率较高。',
                    1: '当前为笔级别信号，需等待线段级别确认以获得更安全的入场时机。',
                    0: ''}.get(signal_level, '')

    # Chan增强
    chan_analysis = ''
    if chan_bonus >= 0.12:
        chan_analysis = '缠论增强系数显著(+0.12以上)，技术面与资金面形成共振。'
    elif chan_bonus >= 0.06:
        chan_analysis = '缠论增强系数中等(+0.06以上)，技术面提供一定支撑。'

    # 因子面
    factor_analysis = ''
    if fn and fn.startswith('DYN_'):
        factor_analysis = '因子面经Walk-Forward IC验证有效，动态因子选择机制确认该因子在当前市场环境中的有效性。'
    elif fn:
        factor_analysis = f'因子面({fn})提供量化支撑，与缠论技术信号形成多维度验证。'
    else:
        factor_analysis = '综合评分体系中各项技术指标协同作用。'

    # 组装段落
    lines = [
        f'【{name}({code})】{industry} | 综合评分: {score:.3f} | {rank_pct_str}',
        '',
        f'  {lesson}分析: {premise}',
        f'  验证: {verification}',
        f'  趋势: {trend_analysis}',
    ]
    if lvl_analysis:
        lines.append(f'  级别: {lvl_analysis}')
    if chan_analysis:
        lines.append(f'  增强: {chan_analysis}')
    lines.append(f'  因子: {factor_analysis}')
    lines.append(f'  结论: {conclusion}，建议纳入候选池。')
    return '\n'.join(lines)


def generate_signal_summary(s):
    """信号强度摘要"""
    buy_pt = s.get("chan_buy_point", 0)
    strength = s.get("chan_buy_strength", 0)
    signal_level = s.get("signal_level", 0)

    if buy_pt > 0 and strength > 0.3:
        lvl_tag = {3: "双级", 2: "段级", 1: "笔级"}.get(signal_level, "")
        return f"Chan{buy_pt}买({lvl_tag}·{strength:.2f})"
    elif buy_pt > 0:
        return f"Chan{buy_pt}买(待确认)"
    else:
        score = s.get("score", 0)
        if score > 0.15: return f"强因子({score:.3f})"
        elif score > 0.05: return f"中因子({score:.3f})"
        else: return f"弱因子({score:.3f})"


def load_portfolio():
    state_file = ROOT / "trade" / "portfolio_state.json"
    if state_file.exists():
        with open(state_file) as f:
            return json.load(f)
    return {"cash": 150000.0, "positions": {}}


# ══════════════════════════════════════════════════════════════
#  选股核心
# ══════════════════════════════════════════════════════════════

def refresh_data():
    """增量更新股票数据 — 确保回测数据可用"""
    print("=" * 60)
    print("  检查数据更新...")
    print("=" * 60)

    from data.data_manager import StockDataManager
    manager = StockDataManager(data_dir=str(ROOT / "data" / "stock_data"))

    # 1. 获取股票列表（网络失败时自动使用缓存）
    stock_list = manager.get_stock_list()
    if len(stock_list) == 0:
        print("股票列表为空，使用现有数据文件直接选股")
        return

    symbols = stock_list['symbol'].tolist()
    symbols.insert(0, 'sh000001')
    print(f"股票池: {len(symbols)} 只")

    # 2. 增量下载原始数据（网络失败时自动使用缓存继续）
    try:
        manager.batch_download(symbols=symbols, force=False, max_workers=2)
    except Exception as e:
        print(f"数据更新异常: {e}")
        print("跳过在线更新，使用本地缓存数据继续...")

    # 3. 同步回测数据: 将 raw 新增日期增量追加到 backtrader 文件
    # create_backtrader_data 内部有增量逻辑，只处理 raw_last_date > bt_last_date 的股票
    from datetime import datetime as dt
    today_str = dt.today().strftime('%Y-%m-%d')
    print("同步回测数据 (仅更新有新增日期的股票)...")
    manager.create_backtrader_data(symbols, start_date='2024-01-01', end_date=today_str, adj_type='qfq')
    print("回测数据已就绪")

    # 4. 下载概念板块当日行情 (供题材热度计算)
    manager.download_concept_daily()
    print()


def run_selection(skip_data_refresh=False):
    """执行选股 — 使用动态因子模式保证精度"""
    cfg = TradeConfig()
    today_str = datetime.today().strftime("%Y-%m-%d")

    print("=" * 60)
    print(f"  选股日期: {today_str}")
    print("=" * 60)

    # ── 先刷新数据 ──
    if skip_data_refresh:
        print("  跳过数据刷新 (已由上游完成)")
    else:
        refresh_data()

    ps = load_portfolio()
    cash = ps.get("cash", 150000.0)
    positions = ps.get("positions", {})
    print(f"当前现金: ¥{cash:,.0f}  持仓: {len(positions)} 只")

    # ── 加载数据 ──
    from trade.signal_runner import SignalRunner
    runner = SignalRunner(
        bt_data_dir=str(cfg.bt_data_dir),
        fund_data_dir=str(cfg.fund_data_dir),
    )
    prices = runner.get_prices()
    print(f"有效价格数据: {len(prices)} 只")

    # ── 正常流程: prepare (含动态因子) + run ──
    runner.prepare(exposure=1.0, peak_equity=0.0)

    # 恢复跨日持仓跟踪状态
    tracking_file = str(ROOT / "trade" / "portfolio_tracking.json")
    if hasattr(runner.strategy, 'portfolio'):
        runner.strategy.portfolio.restore_tracking_state(tracking_file)

    # Fix P1: 从持仓状态提取真实成本, 使止损/移动止盈生效
    cost = {}
    total_market_value = 0.0
    for code, pos in positions.items():
        if isinstance(pos, dict):
            shares = pos.get("shares", 0)
            cost_price = pos.get("cost_price", 0)
            if shares > 0 and cost_price > 0:
                cost[code] = [shares, cost_price]
                total_market_value += shares * prices.get(code, 0)

    # Fix P1: 更新组合日收益率历史 (使波动率控制生效)
    daily_returns = ps.get("daily_returns", [])
    if daily_returns:
        for ret in daily_returns[-60:]:  # 最近60天
            runner.strategy.portfolio.update_returns(ret)

    result = runner.run(
        current_positions={k: v.get("market_value", 0) if isinstance(v, dict) else v
                          for k, v in positions.items()},
        cash=cash,
        cost=cost,
    )

    # 持久化持仓跟踪状态（入场日期/峰值/止损冷却等，跨交易日保持）
    if hasattr(runner.strategy, 'portfolio'):
        runner.strategy.portfolio.save_tracking_state(tracking_file)

    # 输出监控报告
    from strategy.core.monitor import monitor
    monitor.print_report()

    # ── 释放 SignalRunner 中的 stock_data_dict (~4000 只 × 2 年行情, 最大内存占用) ──
    runner.stock_data_dict.clear()
    runner.strategy = None
    del runner
    import gc
    gc.collect()

    if result is None:
        print("选股失败!")
        return None

    return result, prices, cash, today_str


# ══════════════════════════════════════════════════════════════
#  Excel 构建
# ══════════════════════════════════════════════════════════════

def calc_stop_take_profit(s, price):
    """计算缠论结构的止损价和止盈价

    Returns: (stop_price, take_profit_price, risk_reward_ratio)
    """
    sig = s.get("sig")
    buy_pt = s.get("chan_buy_point", 0)

    zg = float(sig.chan_pivot_zg) if sig and not np.isnan(sig.chan_pivot_zg) else 0
    zd = float(sig.chan_pivot_zd) if sig and not np.isnan(sig.chan_pivot_zd) else 0
    zz = float(sig.chan_pivot_zz) if sig and not np.isnan(sig.chan_pivot_zz) else 0
    trend = s.get("trend_type", 0)

    pivot_height = zg - zd if zg > 0 and zd > 0 and zg > zd else price * 0.05

    if buy_pt == 3 and zg > 0:
        # B3: 止损=中枢上沿下方0.5%, 止盈=ZG + 1.5倍中枢高度
        stop = zg * 0.995
        take = zg + pivot_height * 1.5
    elif buy_pt == 2 and zg > 0:
        # B2: 止损=中枢下沿下方0.5%, 止盈=中枢上沿
        stop = zd * 0.995 if zd > 0 else price * 0.95
        take = zg if zg > price else price + pivot_height
    elif buy_pt == 1:
        # B1: 底部反转, 止损宽(8%), 止盈=最近中枢下沿或+12%
        stop = price * 0.92
        take = zd * 1.0 if zd > price else price * 1.12
    elif trend == 2:
        # 上涨趋势无明确买点: 止损=MA20附近(-5%), 止盈=+8%
        stop = price * 0.95
        take = price * 1.08
    elif trend == -2:
        # 下跌趋势: 保守
        stop = price * 0.93
        take = price * 1.06
    else:
        # 默认: 7%止损, 10%止盈
        stop = price * 0.93
        take = price * 1.10

    # 确保止盈>当前价>止损
    stop = min(stop, price * 0.98)
    take = max(take, price * 1.03)

    if price > stop > 0:
        rr = (take - price) / (price - stop)
    else:
        rr = 0

    return round(stop, 2), round(take, 2), round(rr, 2)


def build_excel(result, prices, cash, today_str, stock_names):
    wb = Workbook()
    selections = result.get("selections", [])
    target_positions = result.get("target_positions", {})
    regime = result.get("market_regime", {})
    regime_names = {1: "牛市", 0: "震荡", -1: "熊市"}
    regime_str = regime_names.get(regime.get("regime", 0), "未知")
    momentum = regime.get("momentum_score", 0)
    bear_risk_str = "是" if regime.get("bear_risk") else "否"

    # ─── Sheet 1: 本次选股 ───
    ws1 = wb.active
    ws1.title = "本次选股"

    ws1.merge_cells("A1:O1")
    c = ws1.cell(row=1, column=1, value=f"选股结果 — {today_str}")
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:O2")
    summary = (f"市场状态: {regime_str} | 动量: {momentum:.2f} | "
               f"熊市风险: {bear_risk_str} | 趋势: {regime.get('trend_score', 0):.2f} | "
               f"日期: {today_str}")
    c = ws1.cell(row=2, column=1, value=summary)
    c.font = Font(name="微软雅黑", size=10, color="555555")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22

    headers1 = [
        "序号", "股票代码", "股票名称", "行业", "综合评分",
        "权重(%)", "目标仓位(元)", "当前价格(元)", "目标股数(手)",
        "因子名称", "信号强度", "止损价(元)", "止盈价(元)", "盈亏比", "选股理由"
    ]
    hr = 4
    style_header(ws1, headers1, hr)
    ws1.row_dimensions[hr].height = 22

    if selections:
        for i, s in enumerate(selections):
            code = s.get("code", "")
            row = hr + 1 + i
            rank_pct = s.get("rank_pct", 0)
            rank_str = f"排名前{rank_pct*100:.0f}%" if rank_pct > 0 else ""

            ws1.cell(row=row, column=1, value=i + 1)
            ws1.cell(row=row, column=2, value=code)
            ws1.cell(row=row, column=3, value=stock_names.get(code, ""))
            ws1.cell(row=row, column=4, value=s.get("industry", ""))
            ws1.cell(row=row, column=5, value=round(s.get("score", 0), 4))
            ws1.cell(row=row, column=6, value=round(s.get("weight", 0) * 100, 2))
            target_val = target_positions.get(code, 0)
            ws1.cell(row=row, column=7, value=round(target_val, 0))
            price = prices.get(code, 0)
            ws1.cell(row=row, column=8, value=round(price, 2))
            lots = int(target_val / price / 100) if price > 0 else 0
            ws1.cell(row=row, column=9, value=lots)
            # 因子名称
            fn = s.get("factor_name", "")
            ws1.cell(row=row, column=10, value=fn if fn else "综合因子")
            # 信号强度
            ws1.cell(row=row, column=11, value=generate_signal_summary(s))
            # 止损/止盈/盈亏比
            stop_p, take_p, rr = calc_stop_take_profit(s, price)
            ws1.cell(row=row, column=12, value=stop_p)
            ws1.cell(row=row, column=13, value=take_p)
            ws1.cell(row=row, column=14, value=rr)
            # 选股理由
            ws1.cell(row=row, column=15, value=generate_selection_reason(s, rank_str))

        data_end = hr + len(selections)
        style_data_rows(ws1, hr + 1, data_end, len(headers1))

        total_row = data_end + 1
        total_target = sum(target_positions.get(s.get("code", ""), 0) for s in selections)

        ws1.merge_cells(f"A{total_row}:I{total_row}")
        c = ws1.cell(row=total_row, column=1, value=f"共 {len(selections)} 只股票入选  总仓位: ¥{total_target:,.0f}")
        c.font = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="right", vertical="center")

        for row in range(hr + 1, data_end + 1):
            ws1.cell(row=row, column=5).number_format = "0.0000"
            ws1.cell(row=row, column=6).number_format = "0.00"
            ws1.cell(row=row, column=7).number_format = "#,##0"
            ws1.cell(row=row, column=8).number_format = "0.00"
            ws1.cell(row=row, column=9).number_format = "#,##0"
            ws1.cell(row=row, column=12).number_format = "0.00"
            ws1.cell(row=row, column=13).number_format = "0.00"
            ws1.cell(row=row, column=14).number_format = "0.00"
    else:
        ws1.cell(row=hr + 1, column=1, value="(无选股结果 — 可能筛选条件过严)")

    auto_width(ws1)

    # ─── Sheet 2: 调仓对比 ───
    ws2 = wb.create_sheet("调仓对比")
    ws2.merge_cells("A1:J1")
    c = ws2.cell(row=1, column=1, value=f"调仓对比 — {today_str}")
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers2 = [
        "操作", "股票代码", "股票名称", "行业", "综合评分",
        "权重(%)", "目标仓位(元)", "当前价格(元)", "目标股数(手)", "备注"
    ]
    hr2 = 3
    style_header(ws2, headers2, hr2)

    ps = load_portfolio()
    current_positions = ps.get("positions", {})
    target_codes = {s.get("code") for s in selections}
    current_codes = set(current_positions.keys())
    buy_codes = target_codes - current_codes
    sell_codes = current_codes - target_codes
    hold_codes = target_codes & current_codes

    row_idx = hr2 + 1
    for label, codes, fill in [
        ("买入", buy_codes, GREEN_FILL),
        ("卖出", sell_codes, RED_FILL),
        ("调整", hold_codes, BLUE_FILL),
    ]:
        for code in sorted(codes):
            ws2.cell(row=row_idx, column=1, value=label)
            ws2.cell(row=row_idx, column=2, value=code)
            ws2.cell(row=row_idx, column=3, value=stock_names.get(code, ""))
            sel_info = next((s for s in selections if s.get("code") == code), {})
            ws2.cell(row=row_idx, column=4, value=sel_info.get("industry", ""))
            ws2.cell(row=row_idx, column=5, value=round(sel_info.get("score", 0), 4))
            ws2.cell(row=row_idx, column=6, value=round(sel_info.get("weight", 0) * 100, 2))
            target_val = target_positions.get(code, 0)
            ws2.cell(row=row_idx, column=7, value=round(target_val, 0))
            price = prices.get(code, 0)
            ws2.cell(row=row_idx, column=8, value=round(price, 2))
            lots = int(target_val / price / 100) if price > 0 else 0
            ws2.cell(row=row_idx, column=9, value=lots)
            if label == "卖出":
                pos_info = current_positions.get(code, {})
                ws2.cell(row=row_idx, column=10, value=f"当前持仓: {pos_info}")
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row_idx, column=c).fill = fill
                ws2.cell(row=row_idx, column=c).font = DATA_FONT
                ws2.cell(row=row_idx, column=c).alignment = DATA_ALIGN
                ws2.cell(row=row_idx, column=c).border = THIN_BORDER
            row_idx += 1

    if row_idx == hr2 + 1:
        ws2.cell(row=row_idx, column=1, value="(无调仓)")
        row_idx += 1

    sum_row = row_idx
    ws2.merge_cells(f"A{sum_row}:D{sum_row}")
    ws2.cell(row=sum_row, column=1,
             value=f"买入 {len(buy_codes)} | 卖出 {len(sell_codes)} | "
                   f"调整 {len(hold_codes)} | 共 {len(selections)} 只")
    auto_width(ws2)

    # ─── Sheet 3: 市场状态 ───
    ws3 = wb.create_sheet("市场状态")
    ws3.merge_cells("A1:C1")
    c = ws3.cell(row=1, column=1, value=f"市场状态详情 — {today_str}")
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 30

    headers3 = ["指标", "数值", "说明"]
    hr3 = 3
    style_header(ws3, headers3, hr3)

    market_data = [
        ("日期", today_str, "选股日期"),
        ("市场状态", regime_str, "1=牛市 0=震荡 -1=熊市"),
        ("动量评分", f"{momentum:.4f}", "指数趋势强度，>0.5 强趋势"),
        ("趋势评分", f"{regime.get('trend_score', 0):.4f}", "综合趋势评估"),
        ("熊市风险", bear_risk_str, "是否处于熊市风险区间"),
        ("选股数量", f"{len(selections)} 只", "本次入选股票数"),
        ("目标总仓位", f"¥{sum(target_positions.values()):,.0f}", "所有目标持仓市值之和"),
        ("可用现金", f"¥{cash:,.0f}", "当前可用现金"),
        ("仓位比例", f"{sum(target_positions.values())/cash*100:.1f}%"
         if cash > 0 else "N/A", "目标仓位/现金"),
        ("调仓周期", "20 交易日", "当前调仓频率"),
    ]

    for i, (metric, value, desc) in enumerate(market_data):
        row = hr3 + 1 + i
        ws3.cell(row=row, column=1, value=metric)
        ws3.cell(row=row, column=2, value=value)
        ws3.cell(row=row, column=3, value=desc)

    style_data_rows(ws3, hr3 + 1, hr3 + len(market_data), len(headers3))
    auto_width(ws3)

    # ─── Sheet 4: 调仓历史 ───
    ws4 = wb.create_sheet("调仓历史")
    ws4.merge_cells("A1:K1")
    c = ws4.cell(row=1, column=1, value="调仓历史记录")
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 30

    headers4 = [
        "调仓日期", "操作", "股票代码", "股票名称", "行业",
        "操作前股数", "操作后股数", "成交价(元)", "成交金额(元)", "原因", "备注"
    ]
    hr4 = 3
    style_header(ws4, headers4, hr4)

    if buy_codes:
        for i, code in enumerate(sorted(buy_codes)):
            row = hr4 + 1 + i
            ws4.cell(row=row, column=1, value=today_str)
            ws4.cell(row=row, column=2, value="建仓")
            ws4.cell(row=row, column=3, value=code)
            ws4.cell(row=row, column=4, value=stock_names.get(code, ""))
            sel_info = next((s for s in selections if s.get("code") == code), {})
            ws4.cell(row=row, column=5, value=sel_info.get("industry", ""))
            ws4.cell(row=row, column=6, value=0)
            target_val = target_positions.get(code, 0)
            price = prices.get(code, 0)
            lots = int(target_val / price / 100) if price > 0 else 0
            ws4.cell(row=row, column=7, value=lots * 100)
            ws4.cell(row=row, column=8, value=round(price, 2))
            ws4.cell(row=row, column=9, value=round(target_val, 0))
            ws4.cell(row=row, column=10, value="初始建仓")
        data_end = hr4 + len(buy_codes)
        style_data_rows(ws4, hr4 + 1, data_end, len(headers4))
    else:
        ws4.cell(row=hr4 + 1, column=1, value="(暂无调仓记录)")

    auto_width(ws4)

    # ─── Sheet 5: 信号排名 Top50 ───
    ws5 = wb.create_sheet("信号排名Top50")
    ws5.merge_cells("A1:I1")
    c = ws5.cell(row=1, column=1, value=f"全市场信号排名 Top50 — {today_str}")
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 30

    headers5 = [
        "排名", "股票代码", "股票名称", "行业", "综合评分",
        "信号方向", "因子名称", "当前价格(元)", "选股理由"
    ]
    hr5 = 3
    style_header(ws5, headers5, hr5)

    if selections:
        sorted_sels = sorted(selections, key=lambda x: x.get("score", 0), reverse=True)
        top50 = sorted_sels[:50]
        for i, s in enumerate(top50):
            row = hr5 + 1 + i
            code = s.get("code", "")
            rank_pct = s.get("rank_pct", 0)
            rank_str = f"排名前{rank_pct*100:.0f}%" if rank_pct > 0 else ""
            ws5.cell(row=row, column=1, value=i + 1)
            ws5.cell(row=row, column=2, value=code)
            ws5.cell(row=row, column=3, value=stock_names.get(code, ""))
            ws5.cell(row=row, column=4, value=s.get("industry", ""))
            ws5.cell(row=row, column=5, value=round(s.get("score", 0), 4))
            ws5.cell(row=row, column=6, value="买入")
            fn = s.get("factor_name", "")
            ws5.cell(row=row, column=7, value=fn if fn else "综合因子")
            ws5.cell(row=row, column=8, value=round(prices.get(code, 0), 2))
            ws5.cell(row=row, column=9, value=generate_selection_reason(s, rank_str))

        if len(top50) > 0:
            style_data_rows(ws5, hr5 + 1, hr5 + len(top50), len(headers5))
    else:
        ws5.cell(row=hr5 + 1, column=1, value="(无选股结果)")

    auto_width(ws5)

    # ── 保存 ──
    # ── 按日期保存Excel（不覆盖历史） ──
    date_tag = today_str.replace('-', '')
    output_dir = ROOT / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = str(output_dir / f"选股数据_{date_tag}.xlsx")
    wb.save(output_path)
    print(f"\nExcel 已保存: {output_path}")

    # ── 持仓诊断 ──
    diagnosis = []
    for code, pos in current_positions.items():
        if not isinstance(pos, dict):
            continue
        shares = pos.get('shares', 0)
        cost = pos.get('cost_price', 0)
        live = prices.get(code, 0)
        mv = shares * live if live > 0 else 0
        pnl_pct = ((live - cost) / cost * 100) if cost > 0 and live > 0 else 0

        if code in buy_codes:
            action, advice = '买入', '新建仓位'
        elif code in sell_codes:
            action, advice = ('清仓', '锁定利润') if pnl_pct > 0 else ('清仓', '止损退出')
        elif code in hold_codes:
            if pnl_pct > 5:
                action, advice = '持有', '盈利中，可做T'
            elif pnl_pct > 0:
                action, advice = '持有', '信号仍在，继续持有'
            else:
                action, advice = '持有', '信号仍在，等待反弹'
        else:
            action, advice = ('减仓', '无信号') if pnl_pct > 0 else ('清仓', '无信号且浮亏')

        diagnosis.append({
            'code': code,
            'shares': shares,
            'cost': round(cost, 2),
            'price': round(live, 2),
            'market_value': round(mv, 2),
            'pnl_pct': round(pnl_pct, 2),
            'action': action,
            'advice': advice,
        })

    # ── 输出 QMT target.json ──
    target_json_path = str(ROOT / "trade" / "target.json")
    target_output = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'date': today_str,
        'cash': cash,
        'market_regime': result.get('market_regime', {}),
        'target_positions': result.get('target_positions', {}),
        'buys': [{
            'code': s.get('code', ''),
            'weight': s.get('weight', 0),
            'target_value': result.get('target_positions', {}).get(s.get('code', ''), 0),
            'price': prices.get(s.get('code', ''), 0),
            'score': s.get('score', 0),
            'industry': s.get('industry', ''),
            'chan_buy_point': s.get('chan_buy_point', 0),
            'factor_name': s.get('factor_name', ''),
        } for s in selections],
        'sells': [{
            'code': code,
            'shares': current_positions[code].get('shares', 0) if isinstance(current_positions.get(code), dict) else 0,
            'reason': '不在今日目标中',
        } for code in sell_codes],
        'diagnosis': diagnosis,
        'current_positions': {code: {
            'shares': pos.get('shares', 0) if isinstance(pos, dict) else 0,
            'cost': pos.get('cost_price', 0) if isinstance(pos, dict) else 0,
            'market_value': (pos.get('shares', 0) if isinstance(pos, dict) else 0) * prices.get(code, 0),
        } for code, pos in current_positions.items()},
        'selections': selections,
        'prices': prices,
    }
    os.makedirs(os.path.dirname(target_json_path), exist_ok=True)
    with open(target_json_path, 'w', encoding='utf-8') as f:
        json.dump(target_output, f, ensure_ascii=False, indent=2, default=str)
    print(f"QMT target.json 已生成: {target_json_path}")

    # ── 追加到选股历史CSV（增量记录，不覆盖） ──
    history_path = output_dir / "选股历史.csv"
    history_rows = []
    for s in selections:
        code = s.get("code", "")
        price = prices.get(code, 0)
        target_val = target_positions.get(code, 0)
        stop_p, take_p, rr = calc_stop_take_profit(s, price)
        history_rows.append({
            "选股日期": today_str,
            "股票代码": code,
            "股票名称": stock_names.get(code, ""),
            "行业": s.get("industry", ""),
            "综合评分": round(s.get("score", 0), 4),
            "权重": round(s.get("weight", 0), 4),
            "目标仓位": round(target_val, 0),
            "当前价格": round(price, 2),
            "因子名称": s.get("factor_name", ""),
            "信号强度": generate_signal_summary(s),
            "止损价": stop_p,
            "止盈价": take_p,
            "盈亏比": rr,
            "选股理由": generate_selection_reason(s, f"前{s.get('rank_pct',0)*100:.0f}%" if s.get('rank_pct', 0) > 0 else ""),
        })
    if history_rows:
        df_hist = pd.DataFrame(history_rows)
        if history_path.exists():
            # 增量追加，去重同一天同一股票的记录
            df_old = pd.read_csv(history_path, dtype={"股票代码": str})
            df_combined = pd.concat([df_old, df_hist], ignore_index=True)
            df_combined = df_combined.drop_duplicates(subset=["选股日期", "股票代码"], keep="last")
        else:
            df_combined = df_hist
        df_combined.to_csv(history_path, index=False, encoding="utf-8-sig")
        print(f"选股历史已更新: {history_path} (累计{len(df_combined)}条)")

    return output_path


# ══════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description='选股结果导出到 Excel')
    parser.add_argument('--skip-data-refresh', action='store_true', default=False,
                        help='跳过数据刷新（当上游已执行数据更新时使用）')
    args = parser.parse_args()

    print("加载股票名称...")
    stock_names = load_stock_name_map()
    print(f"股票名称映射: {len(stock_names)} 条")

    result = run_selection(skip_data_refresh=args.skip_data_refresh)
    if result is None:
        return

    result_data, prices, cash, today_str = result

    # runner 需要传给 build_excel 用于 Sheet5
    output = build_excel(result_data, prices, cash, today_str, stock_names)

    # 打印摘要
    print("\n" + "=" * 60)
    print("  选股摘要")
    print("=" * 60)
    selections = result_data.get("selections", [])
    target_positions = result_data.get("target_positions", {})
    if selections:
        for i, s in enumerate(selections):
            code = s.get("code", "")
            name = stock_names.get(code, "")
            price = prices.get(code, 0)
            target = target_positions.get(code, 0)
            lots = int(target / price / 100) if price > 0 else 0
            print(f"  {i+1:2d}. {code} {name:<8s}  {s.get('industry',''):<12s}  "
                  f"评分{s.get('score',0):.4f}  权重{s.get('weight',0)*100:.1f}%  "
                  f"¥{target:>10,.0f}  {lots:>4d}手")
    else:
        print("  (无选股结果)")
        total_signal_store = 0
        # 检查signal_store
        try:
            from trade.signal_runner import SignalRunner
        except Exception:
            pass


if __name__ == "__main__":
    main()
